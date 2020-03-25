# #Create an Excel file
# #Import xlsxwriter module
import xlsxwriter
#
data9 = xlsxwriter.Workbook("data9.xlsx")
# #Creating a workbook using bulletin method and assigning a name to file
worksheet = data9.add_worksheet("Trainees")
worksheet.name = "Data 9 new save"
# #Creating an object
#
# #Assigning string data9 to cell A1
# #
worksheet.write("A1", "First Name")
worksheet.write("B1", "Last Name")
worksheet.write("A2", "Ben")
worksheet.write("B2", "Ellis")
worksheet.write("A3", "Andrew")
worksheet.write("B3", "Wakefield")
worksheet.write("A4", "Jonathan")
worksheet.write("B4", "Williams")
worksheet.write("A5", "CJ")
worksheet.write("B5", "DeGuzman")
worksheet.write("A6", "GiGi")
worksheet.write("B6", "Genius")
worksheet.write("A7", "Weiyee")
worksheet.write("B7", "Lee")
worksheet.write("A8", "Sassy")
worksheet.write("A9", "Shuggs")
worksheet.write("A10", "Tosin")
worksheet.write("A11", "Vijay")
worksheet.write("A12", "Khanhi")
worksheet.write("A13", "Yin")
worksheet.write("A14", "Vivek")
# #Using the bulletin write method we inserted data to our file
#
# data9.close()
# #We used close method to save and close our file


import openpyxl

path = "data9.xlsx"
#Give the location of our file

data9_object = openpyxl.load_workbook(path)

data9_object = data9_object.active

get_data = data9_object.cell(row=1, column=1)

data9_object.cell(row=1, column=1).value = "changed"

print(get_data.value)

total_row = data9_object.max_row
total_column = data9_object.max_column

print(total_row)
print(total_column)

data9_object.title = "New_Sheet_Title"
data9_sheet = data9_object.title

data9.close()



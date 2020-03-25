

#
# import openpyxl
import xlsxwriter

newfile = xlsxwriter.Workbook("newfile.xlsx")

# path1 = "data9.xlsx"
# path2 = "newfile.xlsx"
#
# wb1 = openpyxl.load_workbook(filename=path1)
# ws1 = wb1.worksheets[1]
#
# wb2 = openpyxl.load_workbook(filename=path2)
# ws2 = wb2.create_sheet(ws1,"title")
#
# for row in ws1:
#     for cell in row:
#         ws2[cell.coordinate].value=cell.value
#
# wb2.save(path2)

# importing openpyxl module
import openpyxl as xl;

# opening the source excel file
filename = "C:\\Users\\Andrew Wakefield\\untitled\\Python\\data9.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file
filename1 = "C:\\Users\\Andrew Wakefield\\untitled\\Python\\newfile.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        ws2.cell(row=i, column=j).value = c.value

    # saving the destination excel file
wb2.save(str(filename1))